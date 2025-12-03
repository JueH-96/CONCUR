#!/usr/bin/env python3
import json
import os
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter


def classify_result(jpf_result, jpf_details):
    if not jpf_result or (isinstance(jpf_result, list) and len(jpf_result) == 0):
        return "Termination Error"

    first_result = jpf_result[0] if isinstance(jpf_result, list) else jpf_result

    first_detail = ""
    if jpf_details:
        if isinstance(jpf_details, list) and len(jpf_details) > 0:
            first_detail = jpf_details[0]
        elif isinstance(jpf_details, str):
            first_detail = jpf_details

    if "nullpointerexception" in first_result.lower():
        return "Uncaught Exception"

    if "syntax error" in first_result.lower():
        return "Compilation Errors"

    if "no static entry method" in first_result.lower():
        return "No Entry Method"

    if "deadlock" in first_result.lower():
        return "Deadlock"

    has_results = "====================================================== results" in first_result or \
                  "====================================================== results" in first_detail

    if not has_results:
        return "Termination Error"

    if "preciseracedetector" in first_result.lower():
        if first_detail and "waiting proportion: 100.00%" in first_detail.lower():
            return "Starvation"
        return "Race Condition"

    if "no errors detected" in first_result.lower():
        if first_detail:
            if "unique logical threads created during execution: 1" in first_detail.lower():
                return "Single Thread"
            if "waiting proportion: 100.00%" in first_detail.lower():
                return "Starvation"
        return "Pass"

    return "Uncaught Exception"


def analyze_code_metrics(ground_truth_code):
    """
    Analyze basic metrics of the ground truth code.
    Returns number of lines, number of functions, and number of classes.
    """
    if not ground_truth_code:
        return {
            'lines_of_code': 0,
            'num_functions': 0,
            'num_classes': 0
        }

    lines = ground_truth_code.split('\n')
    loc = 0
    in_multiline_comment = False

    for line in lines:
        stripped = line.strip()
        if not stripped:
            continue
        if '/*' in stripped:
            in_multiline_comment = True
        if in_multiline_comment:
            if '*/' in stripped:
                in_multiline_comment = False
            continue
        if stripped.startswith('//'):
            continue
        loc += 1

    import re
    method_pattern = r'(?:public|private|protected|static|final|\s)+[\w<>\[\]]+\s+\w+\s*\([^)]*\)\s*(?:throws\s+[\w\s,]+)?\s*\{'
    num_functions = len(re.findall(method_pattern, ground_truth_code))

    class_pattern = r'\b(?:public\s+)?(?:abstract\s+)?class\s+\w+'
    num_classes = len(re.findall(class_pattern, ground_truth_code))

    return {
        'lines_of_code': loc,
        'num_functions': num_functions,
        'num_classes': num_classes
    }


def analyze_sync_complexity(ground_truth_code):
    """
    Analyze synchronization complexity of the ground truth code.
    Returns counts of different synchronization primitives.
    """
    if not ground_truth_code:
        return {
            'synchronized': 0,
            'reentrant_lock': 0,
            'lock_interface': 0,
            'semaphore': 0,
            'countdown_latch': 0,
            'cyclic_barrier': 0,
            'read_write_lock': 0,
            'volatile': 0,
            'atomic': 0,
            'blocking_queue': 0,
            'concurrent_hashmap': 0,
            'lock_call': 0,
            'unlock_call': 0,
            'trylock_call': 0,
            'wait_call': 0,
            'notify_call': 0,
            'condition': 0,
            'thread_count': 0,
            'runnable': 0,
            'callable': 0,
            'executor_service': 0
        }

    code = ground_truth_code.lower()

    stats = {
        'synchronized': code.count('synchronized'),
        'reentrant_lock': code.count('reentrantlock'),
        'lock_interface': code.count('lock ') + code.count('lock;') + code.count('lock>'),
        'semaphore': code.count('semaphore'),
        'countdown_latch': code.count('countdownlatch'),
        'cyclic_barrier': code.count('cyclicbarrier'),
        'read_write_lock': code.count('readwritelock') + code.count('reentrantreadwritelock'),
        'volatile': code.count('volatile'),
        'atomic': code.count('atomic'),
        'blocking_queue': code.count('blockingqueue'),
        'concurrent_hashmap': code.count('concurrenthashmap'),
        'lock_call': code.count('.lock()'),
        'unlock_call': code.count('.unlock()'),
        'trylock_call': code.count('.trylock('),
        'wait_call': code.count('.wait()'),
        'notify_call': code.count('.notify()') + code.count('.notifyall()'),
        'condition': code.count('condition'),
        'thread_count': code.count('new thread('),
        'runnable': code.count('runnable'),
        'callable': code.count('callable'),
        'executor_service': code.count('executorservice') + code.count('executors.')
    }

    return stats


def process_json_file(file_path):
    with open(file_path, 'r', encoding='utf-8') as f:
        data = json.load(f)

    results = {
        "Pass": 0,
        "Compilation Errors": 0,
        "Deadlock": 0,
        "Race Condition": 0,
        "Single Thread": 0,
        "Uncaught Exception": 0,
        "No Entry Method": 0,
        "Starvation": 0,
        "Termination Error": 0
    }

    pass1_results = {
        "Pass": 0,
        "Compilation Errors": 0,
        "Deadlock": 0,
        "Race Condition": 0,
        "Single Thread": 0,
        "Uncaught Exception": 0,
        "No Entry Method": 0,
        "Starvation": 0,
        "Termination Error": 0
    }

    error_details = []
    codebleu_pass1 = []
    codebleu_pass3 = []
    pass_at_3_count = 0

    prompt_stats = {}
    prompt_sync_complexity = {}
    prompt_code_metrics = {}

    passed_pass1_codes = []

    for item in data:
        jpf_result = item.get("jpf_result", [])
        jpf_details = item.get("jpf_details", [])
        prompt_title = item.get("prompt_title", "")
        prompt_content = item.get("prompt_content", "") or item.get("prompt", "")
        codebleu_scores = item.get("codebleu_score", [])
        graded_list = item.get("graded_list", [])
        pass_at_3_value = item.get("pass@3", 0)
        ground_truth = item.get("ground_truth", "")
        generated_code = item.get("generated_code", [])
        code_list = item.get("code_list", [])

        if prompt_title and prompt_title not in prompt_sync_complexity:
            prompt_sync_complexity[prompt_title] = analyze_sync_complexity(ground_truth)
            prompt_code_metrics[prompt_title] = analyze_code_metrics(ground_truth)

        classification = classify_result(jpf_result, jpf_details)

        if classification in results:
            results[classification] += 1
        else:
            results["Uncaught Exception"] += 1

        first_jpf_result = [jpf_result[0]] if jpf_result and len(jpf_result) > 0 else []
        first_jpf_details = [jpf_details[0]] if jpf_details and len(jpf_details) > 0 else []
        first_classification = classify_result(first_jpf_result, first_jpf_details)

        if first_classification in pass1_results:
            pass1_results[first_classification] += 1
        else:
            pass1_results["Uncaught Exception"] += 1

        if first_classification == "Pass":
            first_code = generated_code[0] if generated_code and len(generated_code) > 0 else ""
            first_codebleu = codebleu_scores[0] if codebleu_scores and len(codebleu_scores) > 0 else 0
            first_source_code = code_list[0] if code_list and len(code_list) > 0 else ""
            passed_pass1_codes.append({
                "prompt_title": prompt_title,
                "prompt_content": prompt_content,
                "code": first_code,
                "codebleu": first_codebleu,
                "ground_truth": ground_truth,
                "source_code": first_source_code
            })

        if classification != "Pass":
            error_details.append({
                "prompt_title": prompt_title,
                "classification": classification,
                "jpf_result": jpf_result[0] if jpf_result and len(jpf_result) > 0 else "",
                "jpf_details": jpf_details[0] if jpf_details and len(jpf_details) > 0 else ""
            })

        pass_at_3 = False
        pass_at_3_actual_count = 0

        if graded_list and len(graded_list) >= 3:
            pass_at_3_actual_count = sum(1 for x in graded_list[:3] if x == True)
            if pass_at_3_actual_count > 0:
                pass_at_3 = True
                pass_at_3_count += 1
        elif pass_at_3_value > 0:
            pass_at_3_count += 1
            pass_at_3 = True
            pass_at_3_actual_count = 1

        if codebleu_scores and len(codebleu_scores) > 0:
            codebleu_pass1.append(codebleu_scores[0])

        if codebleu_scores and len(codebleu_scores) > 0:
            scores_to_avg = codebleu_scores[:3]
            if scores_to_avg:
                codebleu_pass3.append(sum(scores_to_avg) / len(scores_to_avg))

        if prompt_title not in prompt_stats:
            prompt_stats[prompt_title] = {
                'pass_at_1': 0,
                'pass_at_3': 0,
                'pass_at_3_actual': 0,
                'total': 0
            }

        prompt_stats[prompt_title]['total'] += 1
        if classification == "Pass":
            prompt_stats[prompt_title]['pass_at_1'] += 1
        if pass_at_3:
            prompt_stats[prompt_title]['pass_at_3'] += 1
        prompt_stats[prompt_title]['pass_at_3_actual'] += pass_at_3_actual_count

    return results, pass1_results, error_details, codebleu_pass1, codebleu_pass3, pass_at_3_count, prompt_stats, prompt_sync_complexity, prompt_code_metrics, passed_pass1_codes


def main():
    base_dir = Path("Generated_Code")

    if not base_dir.exists():
        print(f"Error: Directory {base_dir} not found")
        print(f"Please ensure the Generated_Code folder exists")
        return

    all_results = {}
    all_pass1_results = {}
    all_error_details = {}
    all_codebleu_pass1 = {}
    all_codebleu_pass3 = {}
    all_pass_at_3_count = {}
    all_prompt_stats = {}
    all_prompt_sync_complexity = {}
    all_prompt_code_metrics = {}
    all_passed_pass1_codes = {}
    total_items = 0

    llm_dirs = sorted([d for d in base_dir.iterdir() if d.is_dir()])

    if not llm_dirs:
        print(f"Warning: No subdirectories found under {base_dir}")
        return

    print(f"Found {len(llm_dirs)} model directories\n")

    for llm_dir in llm_dirs:
        json_file = llm_dir / "generated_code.json"
        if json_file.exists():
            llm_name = llm_dir.name
            try:
                results, pass1_results, error_details, codebleu_pass1, codebleu_pass3, pass_at_3_count, prompt_stats, prompt_sync_complexity, prompt_code_metrics, passed_pass1_codes = process_json_file(
                    json_file)
                all_results[llm_name] = results
                all_pass1_results[llm_name] = pass1_results
                all_error_details[llm_name] = error_details
                all_codebleu_pass1[llm_name] = codebleu_pass1
                all_codebleu_pass3[llm_name] = codebleu_pass3
                all_pass_at_3_count[llm_name] = pass_at_3_count
                all_prompt_stats[llm_name] = prompt_stats
                all_prompt_sync_complexity[llm_name] = prompt_sync_complexity
                all_prompt_code_metrics[llm_name] = prompt_code_metrics
                all_passed_pass1_codes[llm_name] = passed_pass1_codes
                items_count = sum(results.values())
                total_items += items_count
                print(f"✓ {llm_name}: {items_count} test cases")
            except Exception as e:
                print(f"✗ {llm_name}: Error processing - {str(e)}")
        else:
            print(f"✗ {llm_dir.name}: generated_code.json not found")

    if not all_results:
        print("\nError: No data processed")
        return

    print("\nGenerating Excel report...")

    wb = Workbook()

    sheet1 = wb.active
    sheet1.title = "Pass Rate & CodeBLEU"

    sheet1_headers = ["Model", "Passing Rate (k=1)", "Passing Rate (k=3)", "CodeBLEU (k=1)", "CodeBLEU (k=3)"]
    for col_idx, header in enumerate(sheet1_headers, start=1):
        cell = sheet1.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')

    row_idx = 2
    for llm_name in sorted(all_results.keys()):
        results = all_results[llm_name]
        total_cases = sum(results.values())
        pass_count = results["Pass"]
        pass_at_3_count = all_pass_at_3_count[llm_name]

        pass_k1_str = f"{pass_count}/{total_cases} ({pass_count / total_cases * 100:.2f}%)" if total_cases > 0 else "0/0 (0.00%)"
        pass_k3_str = f"{pass_at_3_count}/{total_cases} ({pass_at_3_count / total_cases * 100:.2f}%)" if total_cases > 0 else "0/0 (0.00%)"

        codebleu_1 = sum(all_codebleu_pass1[llm_name]) / len(all_codebleu_pass1[llm_name]) \
            if all_codebleu_pass1[llm_name] else 0
        codebleu_3 = sum(all_codebleu_pass3[llm_name]) / len(all_codebleu_pass3[llm_name]) \
            if all_codebleu_pass3[llm_name] else 0

        codebleu_1_normalized = codebleu_1 / 100 if codebleu_1 > 1 else codebleu_1
        codebleu_3_normalized = codebleu_3 / 100 if codebleu_3 > 1 else codebleu_3

        sheet1.cell(row=row_idx, column=1, value=llm_name)
        sheet1.cell(row=row_idx, column=2, value=pass_k1_str)
        sheet1.cell(row=row_idx, column=3, value=pass_k3_str)
        sheet1.cell(row=row_idx, column=4, value=round(codebleu_1_normalized, 6))
        sheet1.cell(row=row_idx, column=5, value=round(codebleu_3_normalized, 6))

        row_idx += 1

    for col in range(1, 6):
        sheet1.column_dimensions[get_column_letter(col)].width = 25

    sheet2 = wb.create_sheet("Error Statistics")

    headers = ["Model", "Compilation Errors", "Deadlock (DL)", "Race Condition (RC)",
               "Single Thread (SV)", "Uncaught Exception (UE)", "No Entry Method (NEM)",
               "Starvation (ST)", "Termination Error (TE)"]

    for col_idx, header in enumerate(headers, start=1):
        cell = sheet2.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')

    row_idx = 2
    totals = {key: 0 for key in ["Compilation Errors", "Deadlock", "Race Condition",
                                 "Single Thread", "Uncaught Exception", "No Entry Method",
                                 "Starvation", "Termination Error"]}

    for llm_name, results in sorted(all_results.items()):
        sheet2.cell(row=row_idx, column=1, value=llm_name)
        sheet2.cell(row=row_idx, column=2, value=results["Compilation Errors"])
        sheet2.cell(row=row_idx, column=3, value=results["Deadlock"])
        sheet2.cell(row=row_idx, column=4, value=results["Race Condition"])
        sheet2.cell(row=row_idx, column=5, value=results["Single Thread"])
        sheet2.cell(row=row_idx, column=6, value=results["Uncaught Exception"])
        sheet2.cell(row=row_idx, column=7, value=results["No Entry Method"])
        sheet2.cell(row=row_idx, column=8, value=results["Starvation"])
        sheet2.cell(row=row_idx, column=9, value=results["Termination Error"])

        totals["Compilation Errors"] += results["Compilation Errors"]
        totals["Deadlock"] += results["Deadlock"]
        totals["Race Condition"] += results["Race Condition"]
        totals["Single Thread"] += results["Single Thread"]
        totals["Uncaught Exception"] += results["Uncaught Exception"]
        totals["No Entry Method"] += results["No Entry Method"]
        totals["Starvation"] += results["Starvation"]
        totals["Termination Error"] += results["Termination Error"]

        row_idx += 1

    sheet2.cell(row=row_idx, column=1, value="Total").font = Font(bold=True)
    sheet2.cell(row=row_idx, column=2, value=totals["Compilation Errors"]).font = Font(bold=True)
    sheet2.cell(row=row_idx, column=3, value=totals["Deadlock"]).font = Font(bold=True)
    sheet2.cell(row=row_idx, column=4, value=totals["Race Condition"]).font = Font(bold=True)
    sheet2.cell(row=row_idx, column=5, value=totals["Single Thread"]).font = Font(bold=True)
    sheet2.cell(row=row_idx, column=6, value=totals["Uncaught Exception"]).font = Font(bold=True)
    sheet2.cell(row=row_idx, column=7, value=totals["No Entry Method"]).font = Font(bold=True)
    sheet2.cell(row=row_idx, column=8, value=totals["Starvation"]).font = Font(bold=True)
    sheet2.cell(row=row_idx, column=9, value=totals["Termination Error"]).font = Font(bold=True)

    for col in range(1, 10):
        sheet2.column_dimensions[get_column_letter(col)].width = 20

    sheet3 = wb.create_sheet("Error Details")

    sheet3_headers = ["LLM", "Prompt Title", "Error Type", "JPF Result", "JPF Details"]
    for col_idx, header in enumerate(sheet3_headers, start=1):
        cell = sheet3.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')

    row_idx = 2
    for llm_name in sorted(all_error_details.keys()):
        for error in all_error_details[llm_name]:
            sheet3.cell(row=row_idx, column=1, value=llm_name)
            sheet3.cell(row=row_idx, column=2, value=error["prompt_title"])
            sheet3.cell(row=row_idx, column=3, value=error["classification"])
            sheet3.cell(row=row_idx, column=4, value=error["jpf_result"])
            sheet3.cell(row=row_idx, column=5, value=error["jpf_details"])
            row_idx += 1

    sheet3.column_dimensions['A'].width = 20
    sheet3.column_dimensions['B'].width = 50
    sheet3.column_dimensions['C'].width = 20
    sheet3.column_dimensions['D'].width = 60
    sheet3.column_dimensions['E'].width = 60

    sheet4 = wb.create_sheet("Prompt Pass@1 Rate")

    all_prompts = set()
    for llm_name in all_prompt_stats:
        all_prompts.update(all_prompt_stats[llm_name].keys())
    all_prompts = sorted(all_prompts)

    llm_names = sorted(all_prompt_stats.keys())
    sheet4.cell(row=1, column=1, value="Prompt Title").font = Font(bold=True)
    sheet4.cell(row=1, column=2, value="Total").font = Font(bold=True)
    sheet4.cell(row=1, column=3, value="Pass Count").font = Font(bold=True)
    sheet4.cell(row=1, column=4, value="Lines of Code").font = Font(bold=True)
    sheet4.cell(row=1, column=5, value="Number of Functions").font = Font(bold=True)
    sheet4.cell(row=1, column=6, value="Number of Classes").font = Font(bold=True)

    for col_idx, llm_name in enumerate(llm_names, start=7):
        cell = sheet4.cell(row=1, column=col_idx, value=llm_name)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')

    row_idx = 2
    for prompt in all_prompts:
        sheet4.cell(row=row_idx, column=1, value=prompt)

        total_pass = 0
        total_count = 0
        for llm_name in llm_names:
            if prompt in all_prompt_stats[llm_name]:
                stats = all_prompt_stats[llm_name][prompt]
                total_pass += stats['pass_at_1']
                total_count += stats['total']

        if total_count > 0:
            total_rate = (total_pass / total_count * 100)
            total_str = f"{total_pass}/{total_count} ({total_rate:.2f}%)"
        else:
            total_str = "0/0 (0.00%)"
        sheet4.cell(row=row_idx, column=2, value=total_str)

        sheet4.cell(row=row_idx, column=3, value=total_pass)

        if llm_names and llm_names[0] in all_prompt_code_metrics:
            if prompt in all_prompt_code_metrics[llm_names[0]]:
                metrics = all_prompt_code_metrics[llm_names[0]][prompt]
                sheet4.cell(row=row_idx, column=4, value=metrics['lines_of_code'])
                sheet4.cell(row=row_idx, column=5, value=metrics['num_functions'])
                sheet4.cell(row=row_idx, column=6, value=metrics['num_classes'])
            else:
                sheet4.cell(row=row_idx, column=4, value=0)
                sheet4.cell(row=row_idx, column=5, value=0)
                sheet4.cell(row=row_idx, column=6, value=0)
        else:
            sheet4.cell(row=row_idx, column=4, value=0)
            sheet4.cell(row=row_idx, column=5, value=0)
            sheet4.cell(row=row_idx, column=6, value=0)

        for col_idx, llm_name in enumerate(llm_names, start=7):
            if prompt in all_prompt_stats[llm_name]:
                stats = all_prompt_stats[llm_name][prompt]
                pass_count = stats['pass_at_1']
                total = stats['total']
                rate = (pass_count / total * 100) if total > 0 else 0
                value_str = f"{pass_count}/{total} ({rate:.2f}%)"
                sheet4.cell(row=row_idx, column=col_idx, value=value_str)
            else:
                sheet4.cell(row=row_idx, column=col_idx, value="N/A")

        row_idx += 1

    sheet4.column_dimensions['A'].width = 60
    sheet4.column_dimensions['B'].width = 20
    sheet4.column_dimensions['C'].width = 15
    sheet4.column_dimensions['D'].width = 15
    sheet4.column_dimensions['E'].width = 20
    sheet4.column_dimensions['F'].width = 20
    for col_idx in range(7, len(llm_names) + 7):
        sheet4.column_dimensions[get_column_letter(col_idx)].width = 20

    sheet5 = wb.create_sheet("Prompt Pass@3 Rate")

    sheet5.cell(row=1, column=1, value="Prompt Title").font = Font(bold=True)
    sheet5.cell(row=1, column=2, value="Total").font = Font(bold=True)
    sheet5.cell(row=1, column=3, value="Pass Count").font = Font(bold=True)
    sheet5.cell(row=1, column=4, value="Lines of Code").font = Font(bold=True)
    sheet5.cell(row=1, column=5, value="Number of Functions").font = Font(bold=True)
    sheet5.cell(row=1, column=6, value="Number of Classes").font = Font(bold=True)

    for col_idx, llm_name in enumerate(llm_names, start=7):
        cell = sheet5.cell(row=1, column=col_idx, value=llm_name)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')

    row_idx = 2
    for prompt in all_prompts:
        sheet5.cell(row=row_idx, column=1, value=prompt)

        total_pass_actual = 0
        total_attempts = 0

        for llm_name in llm_names:
            if prompt in all_prompt_stats[llm_name]:
                stats = all_prompt_stats[llm_name][prompt]
                total_pass_actual += stats['pass_at_3_actual']
                total_attempts += stats['total'] * 3

        if total_attempts > 0:
            total_rate = (total_pass_actual / total_attempts * 100)
            total_str = f"{total_pass_actual}/{total_attempts} ({total_rate:.2f}%)"
        else:
            total_str = "0/0 (0.00%)"
        sheet5.cell(row=row_idx, column=2, value=total_str)

        sheet5.cell(row=row_idx, column=3, value=total_pass_actual)

        if llm_names and llm_names[0] in all_prompt_code_metrics:
            if prompt in all_prompt_code_metrics[llm_names[0]]:
                metrics = all_prompt_code_metrics[llm_names[0]][prompt]
                sheet5.cell(row=row_idx, column=4, value=metrics['lines_of_code'])
                sheet5.cell(row=row_idx, column=5, value=metrics['num_functions'])
                sheet5.cell(row=row_idx, column=6, value=metrics['num_classes'])
            else:
                sheet5.cell(row=row_idx, column=4, value=0)
                sheet5.cell(row=row_idx, column=5, value=0)
                sheet5.cell(row=row_idx, column=6, value=0)
        else:
            sheet5.cell(row=row_idx, column=4, value=0)
            sheet5.cell(row=row_idx, column=5, value=0)
            sheet5.cell(row=row_idx, column=6, value=0)

        for col_idx, llm_name in enumerate(llm_names, start=7):
            if prompt in all_prompt_stats[llm_name]:
                stats = all_prompt_stats[llm_name][prompt]
                pass_actual = stats['pass_at_3_actual']
                total_attempts_llm = stats['total'] * 3
                rate = (pass_actual / total_attempts_llm * 100) if total_attempts_llm > 0 else 0
                value_str = f"{pass_actual}/{total_attempts_llm} ({rate:.2f}%)"
                sheet5.cell(row=row_idx, column=col_idx, value=value_str)
            else:
                sheet5.cell(row=row_idx, column=col_idx, value="N/A")

        row_idx += 1

    sheet5.column_dimensions['A'].width = 60
    sheet5.column_dimensions['B'].width = 20
    sheet5.column_dimensions['C'].width = 15
    sheet5.column_dimensions['D'].width = 15
    sheet5.column_dimensions['E'].width = 20
    sheet5.column_dimensions['F'].width = 20
    for col_idx in range(7, len(llm_names) + 7):
        sheet5.column_dimensions[get_column_letter(col_idx)].width = 20

    sheet6 = wb.create_sheet("Sync Complexity")

    first_llm = llm_names[0] if llm_names else None
    if first_llm and first_llm in all_prompt_sync_complexity:
        complexity_data = all_prompt_sync_complexity[first_llm]
    else:
        complexity_data = {}

    headers = [
        "Prompt Title",
        "synchronized",
        "ReentrantLock",
        "Lock Interface",
        "Semaphore",
        "CountDownLatch",
        "CyclicBarrier",
        "ReadWriteLock",
        "volatile",
        "Atomic*",
        "BlockingQueue",
        "ConcurrentHashMap",
        "lock()",
        "unlock()",
        "tryLock()",
        "wait()",
        "notify/notifyAll",
        "Condition",
        "Thread Count",
        "Runnable",
        "Callable",
        "ExecutorService"
    ]

    for col_idx, header in enumerate(headers, start=1):
        cell = sheet6.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')

    row_idx = 2
    for prompt in sorted(complexity_data.keys()):
        stats = complexity_data[prompt]
        sheet6.cell(row=row_idx, column=1, value=prompt)
        sheet6.cell(row=row_idx, column=2, value=stats['synchronized'])
        sheet6.cell(row=row_idx, column=3, value=stats['reentrant_lock'])
        sheet6.cell(row=row_idx, column=4, value=stats['lock_interface'])
        sheet6.cell(row=row_idx, column=5, value=stats['semaphore'])
        sheet6.cell(row=row_idx, column=6, value=stats['countdown_latch'])
        sheet6.cell(row=row_idx, column=7, value=stats['cyclic_barrier'])
        sheet6.cell(row=row_idx, column=8, value=stats['read_write_lock'])
        sheet6.cell(row=row_idx, column=9, value=stats['volatile'])
        sheet6.cell(row=row_idx, column=10, value=stats['atomic'])
        sheet6.cell(row=row_idx, column=11, value=stats['blocking_queue'])
        sheet6.cell(row=row_idx, column=12, value=stats['concurrent_hashmap'])
        sheet6.cell(row=row_idx, column=13, value=stats['lock_call'])
        sheet6.cell(row=row_idx, column=14, value=stats['unlock_call'])
        sheet6.cell(row=row_idx, column=15, value=stats['trylock_call'])
        sheet6.cell(row=row_idx, column=16, value=stats['wait_call'])
        sheet6.cell(row=row_idx, column=17, value=stats['notify_call'])
        sheet6.cell(row=row_idx, column=18, value=stats['condition'])
        sheet6.cell(row=row_idx, column=19, value=stats['thread_count'])
        sheet6.cell(row=row_idx, column=20, value=stats['runnable'])
        sheet6.cell(row=row_idx, column=21, value=stats['callable'])
        sheet6.cell(row=row_idx, column=22, value=stats['executor_service'])
        row_idx += 1

    sheet6.column_dimensions['A'].width = 60
    for col_idx in range(2, 23):
        sheet6.column_dimensions[get_column_letter(col_idx)].width = 15

    sheet7 = wb.create_sheet("Error Statistics (Pass@1)")

    sheet7_headers = ["LLM", "Compilation Errors", "Deadlock", "Race Condition", "Single Thread",
                      "Uncaught Exception", "No Entry Method", "Starvation", "Termination Error"]
    for col_idx, header in enumerate(sheet7_headers, start=1):
        cell = sheet7.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')

    row_idx = 2
    totals_pass1 = {
        "Compilation Errors": 0,
        "Deadlock": 0,
        "Race Condition": 0,
        "Single Thread": 0,
        "Uncaught Exception": 0,
        "No Entry Method": 0,
        "Starvation": 0,
        "Termination Error": 0
    }

    for llm_name in sorted(all_pass1_results.keys()):
        pass1_results = all_pass1_results[llm_name]
        sheet7.cell(row=row_idx, column=1, value=llm_name)
        sheet7.cell(row=row_idx, column=2, value=pass1_results["Compilation Errors"])
        sheet7.cell(row=row_idx, column=3, value=pass1_results["Deadlock"])
        sheet7.cell(row=row_idx, column=4, value=pass1_results["Race Condition"])
        sheet7.cell(row=row_idx, column=5, value=pass1_results["Single Thread"])
        sheet7.cell(row=row_idx, column=6, value=pass1_results["Uncaught Exception"])
        sheet7.cell(row=row_idx, column=7, value=pass1_results["No Entry Method"])
        sheet7.cell(row=row_idx, column=8, value=pass1_results["Starvation"])
        sheet7.cell(row=row_idx, column=9, value=pass1_results["Termination Error"])

        totals_pass1["Compilation Errors"] += pass1_results["Compilation Errors"]
        totals_pass1["Deadlock"] += pass1_results["Deadlock"]
        totals_pass1["Race Condition"] += pass1_results["Race Condition"]
        totals_pass1["Single Thread"] += pass1_results["Single Thread"]
        totals_pass1["Uncaught Exception"] += pass1_results["Uncaught Exception"]
        totals_pass1["No Entry Method"] += pass1_results["No Entry Method"]
        totals_pass1["Starvation"] += pass1_results["Starvation"]
        totals_pass1["Termination Error"] += pass1_results["Termination Error"]

        row_idx += 1

    sheet7.cell(row=row_idx, column=1, value="Total").font = Font(bold=True)
    sheet7.cell(row=row_idx, column=2, value=totals_pass1["Compilation Errors"]).font = Font(bold=True)
    sheet7.cell(row=row_idx, column=3, value=totals_pass1["Deadlock"]).font = Font(bold=True)
    sheet7.cell(row=row_idx, column=4, value=totals_pass1["Race Condition"]).font = Font(bold=True)
    sheet7.cell(row=row_idx, column=5, value=totals_pass1["Single Thread"]).font = Font(bold=True)
    sheet7.cell(row=row_idx, column=6, value=totals_pass1["Uncaught Exception"]).font = Font(bold=True)
    sheet7.cell(row=row_idx, column=7, value=totals_pass1["No Entry Method"]).font = Font(bold=True)
    sheet7.cell(row=row_idx, column=8, value=totals_pass1["Starvation"]).font = Font(bold=True)
    sheet7.cell(row=row_idx, column=9, value=totals_pass1["Termination Error"]).font = Font(bold=True)

    for col in range(1, 10):
        sheet7.column_dimensions[get_column_letter(col)].width = 20

    sheet8 = wb.create_sheet("Passed Pass@1 Codes")

    all_passed_codes_list = []
    for llm_name in sorted(all_passed_pass1_codes.keys()):
        for code_info in all_passed_pass1_codes[llm_name]:
            all_passed_codes_list.append({
                "llm": llm_name,
                "prompt_title": code_info["prompt_title"],
                "prompt_content": code_info["prompt_content"],
                "code": code_info["code"],
                "codebleu": code_info["codebleu"],
                "ground_truth": code_info["ground_truth"],
                "source_code": code_info["source_code"]
            })

    sheet8_headers = ["LLM", "Prompt Title", "Prompt Content", "CodeBLEU", "Generated Code", "Ground Truth",
                      "Source Code"]
    for col_idx, header in enumerate(sheet8_headers, start=1):
        cell = sheet8.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')

    row_idx = 2
    for code_info in all_passed_codes_list:
        sheet8.cell(row=row_idx, column=1, value=code_info["llm"])
        sheet8.cell(row=row_idx, column=2, value=code_info["prompt_title"])
        sheet8.cell(row=row_idx, column=3, value=code_info["prompt_content"])
        sheet8.cell(row=row_idx, column=4, value=code_info["codebleu"])
        sheet8.cell(row=row_idx, column=5, value=code_info["code"])
        sheet8.cell(row=row_idx, column=6, value=code_info["ground_truth"])
        sheet8.cell(row=row_idx, column=7, value=code_info["source_code"])
        row_idx += 1

    sheet8.column_dimensions['A'].width = 20
    sheet8.column_dimensions['B'].width = 50
    sheet8.column_dimensions['C'].width = 60
    sheet8.column_dimensions['D'].width = 15
    sheet8.column_dimensions['E'].width = 80
    sheet8.column_dimensions['F'].width = 80
    sheet8.column_dimensions['G'].width = 80

    sheet9 = wb.create_sheet("Selected 115 Codes")

    prompt_code_groups = {}
    for code_info in all_passed_codes_list:
        prompt = code_info["prompt_title"]
        if prompt not in prompt_code_groups:
            prompt_code_groups[prompt] = []
        prompt_code_groups[prompt].append(code_info)

    selected_codes = []
    target_count = 115

    if len(prompt_code_groups) > 0:
        sorted_prompts = sorted(prompt_code_groups.keys())

        for prompt in sorted_prompts:
            codes = prompt_code_groups[prompt]
            codes_sorted = sorted(codes, key=lambda x: x["codebleu"], reverse=True)
            selected_codes.append(codes_sorted[0])

            if len(selected_codes) >= target_count:
                break

        if len(selected_codes) < target_count:
            round_num = 2
            while len(selected_codes) < target_count:
                added_in_round = False
                for prompt in sorted_prompts:
                    if len(selected_codes) >= target_count:
                        break
                    codes = prompt_code_groups[prompt]
                    codes_sorted = sorted(codes, key=lambda x: x["codebleu"], reverse=True)
                    if len(codes_sorted) >= round_num:
                        selected_codes.append(codes_sorted[round_num - 1])
                        added_in_round = True
                if not added_in_round:
                    break
                round_num += 1

        selected_codes = selected_codes[:target_count]

    sheet9_headers = ["LLM", "Prompt Title", "Prompt Content", "CodeBLEU", "Generated Code", "Ground Truth",
                      "Source Code"]
    for col_idx, header in enumerate(sheet9_headers, start=1):
        cell = sheet9.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')

    row_idx = 2
    for code_info in selected_codes:
        sheet9.cell(row=row_idx, column=1, value=code_info["llm"])
        sheet9.cell(row=row_idx, column=2, value=code_info["prompt_title"])
        sheet9.cell(row=row_idx, column=3, value=code_info["prompt_content"])
        sheet9.cell(row=row_idx, column=4, value=code_info["codebleu"])
        sheet9.cell(row=row_idx, column=5, value=code_info["code"])
        sheet9.cell(row=row_idx, column=6, value=code_info["ground_truth"])
        sheet9.cell(row=row_idx, column=7, value=code_info["source_code"])
        row_idx += 1

    sheet9.column_dimensions['A'].width = 20
    sheet9.column_dimensions['B'].width = 50
    sheet9.column_dimensions['C'].width = 60
    sheet9.column_dimensions['D'].width = 15
    sheet9.column_dimensions['E'].width = 80
    sheet9.column_dimensions['F'].width = 80
    sheet9.column_dimensions['G'].width = 80

    output_file = "code_analysis_results.xlsx"
    wb.save(output_file)

    print(f"\n{'=' * 60}")
    print(f"✓ Analysis completed!")
    print(f"  - Processed {len(all_results)} LLM models")
    print(f"  - Total {total_items} test cases")
    print(f"  - Sheet1 (Pass Rate & CodeBLEU)")
    print(f"  - Sheet2 (Error Statistics)")
    print(f"  - Sheet3 (Error Details)")
    print(f"  - Sheet4 (Prompt Pass@1 Rate)")
    print(f"  - Sheet5 (Prompt Pass@3 Rate)")
    print(f"  - Sheet6 (Sync Complexity)")
    print(f"  - Sheet7 (Error Statistics Pass@1)")
    total_passed_codes = sum(len(codes) for codes in all_passed_pass1_codes.values())
    print(f"  - Sheet8 (Passed Pass@1 Codes): {total_passed_codes} codes")
    print(f"  - Sheet9 (Selected 115 Codes)")
    print(f"  - Saved to: {output_file}")
    print(f"{'=' * 60}")


if __name__ == "__main__":
    main()
